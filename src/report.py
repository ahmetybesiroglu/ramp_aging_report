# report.py

import pandas as pd
import numpy as np
from .utils import convert_to_iso8601, categorize_aging, generate_column_names
from .api import get_ramp_api_token, get_entities, get_bills
from datetime import datetime
from openpyxl import load_workbook
import xlrd
import xml.etree.ElementTree as ET

def filter_open_as_of(df, cut_off_date):
    """
    Filters bills that were still open as of the specified cutoff date.
    Uses 'payment.effective_date' if available, otherwise falls back to 'paid_at'.
    """
    # Extract 'effective_date' from the 'payment' column
    df['effective_date'] = pd.to_datetime(
        df.apply(
            lambda x: x['payment'].get('effective_date') if isinstance(x['payment'], dict) and 'effective_date' in x['payment'] else x['paid_at'],
            axis=1
        ),
        errors='coerce'
    ).dt.tz_localize(None)
    
    # A bill is considered open if it was not paid (no effective_date) or was paid after the cutoff date
    open_as_of_cutoff = df['effective_date'].isna() | (df['effective_date'] > cut_off_date)
    
    # Filter bills that were open as of the cutoff date
    return df.loc[open_as_of_cutoff].copy()



def save_raw_data(entity_name, df, specified_date):
    """
    Saves the raw bills data to a CSV file for each entity.
    """
    raw_filename = f'{entity_name.replace(" ", "_").lower()}_raw_bills_data_as_of_{specified_date}.csv'
    df.to_csv(raw_filename, index=False)
    print(f"Raw data saved for entity {entity_name} as {raw_filename}")

def generate_entity_reports(specified_date):
    """
    Generates aging reports for each individual entity and saves them as CSV files.
    Additionally, saves the raw bills data for each entity.
    """
    # Get the API access token
    access_token = get_ramp_api_token()

    # Convert the user-friendly specified date to ISO 8601 format
    cut_off_date_iso8601 = convert_to_iso8601(specified_date)
    cut_off_date = datetime.strptime(cut_off_date_iso8601, "%Y-%m-%dT%H:%M:%SZ")

    # Fetch all entities
    entities = get_entities(access_token)

    # Loop through each entity and generate the individual report
    for entity in entities:
        entity_id = entity['id']
        entity_name = entity['entity_name']

        # Get bills for the entity (fetches all, no payment status filter)
        bills_data = get_bills(access_token, entity_id, cut_off_date_iso8601)

        if bills_data:
            # Load the bills data into a DataFrame
            df = pd.DataFrame(bills_data)

            # Save the raw data before any filtering or processing
            save_raw_data(entity_name, df, specified_date)

            # Ensure due_at and issued_at are in the correct datetime format
            df['due_at'] = pd.to_datetime(df['due_at'], errors='coerce').dt.tz_localize(None)
            df['issued_at'] = pd.to_datetime(df['issued_at'], errors='coerce').dt.tz_localize(None)

            # Extract amount and vendor information (divide by 100 to correct extra zeros)
            df['amount'] = df['amount'].apply(lambda x: (x['amount'] / 100) if isinstance(x, dict) else (x / 100))
            df['vendor_name'] = df['vendor'].apply(lambda x: x['remote_name'].strip() if isinstance(x, dict) and 'remote_name' in x else x.strip())

            # Filter for bills that were still open as of the cutoff date using updated effective_date logic
            df_open = filter_open_as_of(df, cut_off_date)

            if df_open.empty:
                print(f"No bills were open as of {specified_date} for entity {entity_name}")
                continue

            # Apply categorization based on the cutoff date
            df_open['aging_bucket'] = df_open.apply(lambda row: categorize_aging(row, cut_off_date), axis=1)

            # Group by vendor and aging bucket, and sum the amounts
            aging_report = df_open.groupby(['vendor_name', 'aging_bucket']).agg({'amount': 'sum'}).unstack(fill_value=0).reset_index()

            # Flatten the column multi-index
            aging_report.columns = ['Vendor Name'] + [f'{col[1]}' for col in aging_report.columns[1:]]

            # Calculate the total for each vendor
            aging_report['Total'] = aging_report.iloc[:, 1:].sum(axis=1)

            # Generate dynamic column names based on the cutoff date
            sorted_columns = generate_column_names(cut_off_date)

            # Ensure the columns are ordered as specified
            aging_report = aging_report.reindex(columns=sorted_columns, fill_value=0)

            # Create a new column for sorting in the aging report
            aging_report['Vendor Name Sort'] = aging_report['Vendor Name'].str.lower()

            # Sort rows by the new sorting column and drop it afterwards
            aging_report = aging_report.sort_values(by='Vendor Name Sort').drop(columns=['Vendor Name Sort'])

            # Calculate the sum for each column to get the totals
            totals_row = pd.DataFrame(aging_report.iloc[:, 1:].sum(axis=0)).T
            totals_row['Vendor Name'] = 'Total'

            # Append the totals row to the aging report using pd.concat
            aging_report = pd.concat([aging_report, totals_row], ignore_index=True)

            # Save the aging report as a CSV file for this entity
            filename = f'{entity_name.replace(" ", "_").lower()}_open_bills_aging_report_as_of_{specified_date}.csv'
            aging_report.to_csv(filename, index=False)

            print(f"Aging report generated for entity: {entity_name}")
        else:
            print(f"No bills data for entity {entity_name}")




def generate_combined_report(specified_date):
    """
    Generates a combined aging report for all entities and saves it as a single CSV file.
    Additionally, saves the combined raw bills data for later analysis.
    """
    # Get the API access token
    access_token = get_ramp_api_token()

    # Convert the user-friendly specified date to ISO 8601 format
    cut_off_date_iso8601 = convert_to_iso8601(specified_date)
    cut_off_date = datetime.strptime(cut_off_date_iso8601, "%Y-%m-%dT%H:%M:%SZ")

    # Fetch all entities
    entities = get_entities(access_token)

    # Create an empty DataFrame to store combined data
    combined_bills_data = pd.DataFrame()

    # Loop through each entity and fetch its bills
    for entity in entities:
        entity_id = entity['id']
        entity_name = entity['entity_name']

        # Get bills for the entity (fetches all, no payment status filter)
        bills_data = get_bills(access_token, entity_id, cut_off_date_iso8601)

        if bills_data:
            # Load the bills data into a DataFrame
            df = pd.DataFrame(bills_data)

            # Ensure due_at and issued_at are in the correct datetime format
            df['due_at'] = pd.to_datetime(df['due_at'], errors='coerce').dt.tz_localize(None)
            df['issued_at'] = pd.to_datetime(df['issued_at'], errors='coerce').dt.tz_localize(None)

            # Extract amount and vendor information (divide by 100 to correct extra zeros)
            df['amount'] = df['amount'].apply(lambda x: (x['amount'] / 100) if isinstance(x, dict) else (x / 100))
            df['vendor_name'] = df['vendor'].apply(lambda x: x['remote_name'].strip() if isinstance(x, dict) and 'remote_name' in x else x.strip())

            # Filter for bills that were still open as of the cutoff date using updated effective_date logic
            df_open = filter_open_as_of(df, cut_off_date)

            # Add the filtered DataFrame to the combined DataFrame
            combined_bills_data = pd.concat([combined_bills_data, df_open], ignore_index=True)

    # If no bills data was retrieved, exit early
    if combined_bills_data.empty:
        print("No bills data was retrieved from any entity.")
        return

    # Apply categorization based on the cutoff date
    combined_bills_data['aging_bucket'] = combined_bills_data.apply(lambda row: categorize_aging(row, cut_off_date), axis=1)

    # Group by vendor and aging bucket, and sum the amounts
    aging_report = combined_bills_data.groupby(['vendor_name', 'aging_bucket']).agg({'amount': 'sum'}).unstack(fill_value=0).reset_index()

    # Flatten the column multi-index
    aging_report.columns = ['Vendor Name'] + [f'{col[1]}' for col in aging_report.columns[1:]]

    # Calculate the total for each vendor
    aging_report['Total'] = aging_report.iloc[:, 1:].sum(axis=1)

    # Generate dynamic column names based on the cutoff date
    sorted_columns = generate_column_names(cut_off_date)

    # Ensure the columns are ordered as specified
    aging_report = aging_report.reindex(columns=sorted_columns, fill_value=0)

    # Create a new column for sorting in the aging report
    aging_report['Vendor Name Sort'] = aging_report['Vendor Name'].str.lower()

    # Sort rows by the new sorting column and drop it afterwards
    aging_report = aging_report.sort_values(by='Vendor Name Sort').drop(columns=['Vendor Name Sort'])

    # Calculate the sum for each column to get the totals
    totals_row = pd.DataFrame(aging_report.iloc[:, 1:].sum(axis=0)).T
    totals_row['Vendor Name'] = 'Total'

    # Append the totals row to the aging report using pd.concat
    aging_report = pd.concat([aging_report, totals_row], ignore_index=True)

    # Save the combined aging report as a CSV file
    filename = f'combined_open_bills_aging_report_as_of_{specified_date}.csv'
    aging_report.to_csv(filename, index=False)

    # Print the aging report
    print(f"Combined aging report as of {specified_date}:")
    print(aging_report)


import pandas as pd
import xml.etree.ElementTree as ET

import pandas as pd
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np

import pandas as pd
import xml.etree.ElementTree as ET

def generate_reconciliation_report(ramp_report_path, netsuite_report_path, output_path, epsilon=1e-6):
    # Read Ramp report
    ramp_df = pd.read_csv(ramp_report_path)
    
    # Rename Ramp columns
    ramp_df.columns = ['vendor', 'ramp_current', 'ramp_30', 'ramp_60', 'ramp_90', 'ramp_>90', 'ramp_total']
    
    # Read NetSuite XML-formatted Excel report
    try:
        tree = ET.parse(netsuite_report_path)
        root = tree.getroot()
        
        # Define the XML namespace
        ns = {'ss': 'urn:schemas-microsoft-com:office:spreadsheet'}
        
        # Extract data from XML
        netsuite_data = []
        for row in root.findall('.//ss:Row', ns)[11:]:  # Start from 11th row
            cells = row.findall('ss:Cell', ns)
            row_data = []
            for cell in cells[:7]:  # Take only first 7 columns
                data = cell.find('ss:Data', ns)
                row_data.append(data.text if data is not None else '')
            netsuite_data.append(row_data)
        
        # Convert to DataFrame
        netsuite_df = pd.DataFrame(netsuite_data, columns=['vendor', 'netsuite_current', 'netsuite_30', 'netsuite_60', 'netsuite_90', 'netsuite_>90', 'netsuite_total'])
        
        # Convert numeric columns to float
        numeric_columns = ['netsuite_current', 'netsuite_30', 'netsuite_60', 'netsuite_90', 'netsuite_>90', 'netsuite_total']
        for col in numeric_columns:
            netsuite_df[col] = pd.to_numeric(netsuite_df[col].str.replace('[$,]', '', regex=True), errors='coerce')
        
    except Exception as e:
        print(f"Error reading NetSuite file: {e}")
        return
    
    # Merge dataframes
    merged_df = pd.merge(ramp_df, netsuite_df, on='vendor', how='outer')
    
    # Filter out vendors starting with "IC - " or exactly "Total"
    merged_df = merged_df[~merged_df['vendor'].str.startswith('IC - ') & (merged_df['vendor'] != 'Total')]
    
    # Calculate differences
    for period in ['current', '30', '60', '90', '>90', 'total']:
        merged_df[f'diff_{period}'] = merged_df[f'ramp_{period}'] - merged_df[f'netsuite_{period}']
    
    # Apply epsilon to the diff columns to remove negligible values
    for period in ['current', '30', '60', '90', '>90', 'total']:
        diff_col = f'diff_{period}'
        merged_df[diff_col] = merged_df[diff_col].apply(lambda x: 0 if abs(x) < epsilon else x)
    
    # Reorder columns
    column_order = ['vendor']
    for period in ['current', '30', '60', '90', '>90', 'total']:
        column_order.extend([f'ramp_{period}', f'netsuite_{period}', f'diff_{period}'])
    
    merged_df = merged_df[column_order]
    
    # Replace NaN with 0 for numerical columns
    numeric_columns = merged_df.columns.drop('vendor')
    merged_df[numeric_columns] = merged_df[numeric_columns].fillna(0)
    
    # Sort by vendor name alphabetically
    merged_df = merged_df.sort_values('vendor')
    
    # Calculate totals for each numeric column
    total_row = pd.DataFrame(merged_df[numeric_columns].sum()).transpose()
    total_row['vendor'] = 'Total'
    
    # Append the total row to the DataFrame
    merged_df = pd.concat([merged_df, total_row], ignore_index=True)
    
    # Reorder to make sure 'Total' is the last row
    merged_df = merged_df.reset_index(drop=True)
    
    # Save to CSV
    merged_df.to_csv(output_path, index=False)
    
    print(f"Reconciliation report saved to {output_path}")
    
    return merged_df
